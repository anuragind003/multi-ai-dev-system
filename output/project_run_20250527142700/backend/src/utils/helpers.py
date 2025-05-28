import datetime
import uuid
import io
import csv
from flask import send_file, Response
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

def validate_payload(data: dict, required_fields: list, optional_fields: list = None) -> tuple[bool, str]:
    """
    Validates an incoming dictionary payload against a list of required fields.

    Args:
        data (dict): The dictionary payload to validate.
        required_fields (list): A list of strings representing fields that must be present.
        optional_fields (list, optional): A list of strings representing fields that are optional.
                                          Defaults to None.

    Returns:
        tuple[bool, str]: A tuple where the first element is True if validation passes,
                          False otherwise. The second element is an error message if validation fails.
    """
    if not isinstance(data, dict):
        return False, "Payload must be a JSON object."

    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return False, f"Missing required fields: {', '.join(missing_fields)}"

    return True, "Payload is valid."

def parse_iso_timestamp(timestamp_str: str) -> datetime.datetime | None:
    """
    Parses an ISO 8601 formatted timestamp string into a datetime object.
    Handles various ISO 8601 formats including those with/without timezone and microseconds.

    Args:
        timestamp_str (str): The timestamp string to parse.

    Returns:
        datetime.datetime | None: A datetime object if parsing is successful, None otherwise.
    """
    if not isinstance(timestamp_str, str):
        return None
    try:
        # datetime.fromisoformat handles various ISO 8601 formats
        # e.g., 'YYYY-MM-DDTHH:MM:SS', 'YYYY-MM-DDTHH:MM:SS.ffffff',
        # 'YYYY-MM-DDTHH:MM:SS+HH:MM', 'YYYY-MM-DDTHH:MM:SS.ffffff+HH:MM'
        dt_obj = datetime.datetime.fromisoformat(timestamp_str)
        # If the datetime object is naive (no timezone info), assume UTC for consistency.
        if dt_obj.tzinfo is None:
            return dt_obj.replace(tzinfo=datetime.timezone.utc)
        return dt_obj
    except ValueError:
        return None

def is_valid_uuid(uuid_string: str) -> bool:
    """
    Checks if a string is a valid UUID.

    Args:
        uuid_string (str): The string to check.

    Returns:
        bool: True if the string is a valid UUID, False otherwise.
    """
    if not isinstance(uuid_string, str):
        return False
    try:
        uuid.UUID(uuid_string)
        return True
    except ValueError:
        return False

def generate_csv_response(data: list[dict], filename: str, headers: list[str]) -> Response:
    """
    Generates an in-memory CSV file from a list of dictionaries and returns it as a Flask response.

    Args:
        data (list[dict]): A list of dictionaries, where each dictionary represents a row.
        filename (str): The desired filename for the downloaded CSV.
        headers (list[str]): A list of strings representing the CSV column headers.
                              These should match keys in the data dictionaries.

    Returns:
        Response: A Flask response object containing the CSV file.
    """
    si = io.StringIO()
    cw = csv.writer(si)

    # Write headers
    cw.writerow(headers)

    # Write data rows
    for row_data in data:
        row = [row_data.get(header, '') for header in headers]
        cw.writerow(row)

    output = io.BytesIO(si.getvalue().encode('utf-8'))
    output.seek(0)

    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f"{filename}.csv"
    )

def generate_excel_response(data: list[dict], filename: str, headers: list[str]) -> Response:
    """
    Generates an in-memory Excel (XLSX) file from a list of dictionaries and returns it as a Flask response.

    Args:
        data (list[dict]): A list of dictionaries, where each dictionary represents a row.
        filename (str): The desired filename for the downloaded Excel file.
        headers (list[str]): A list of strings representing the Excel column headers.
                              These should match keys in the data dictionaries.

    Returns:
        Response: A Flask response object containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Apply header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Write headers
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    for row_data in data:
        row = [row_data.get(header, '') for header in headers]
        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{filename}.xlsx"
    )

def is_customer_eligible_for_campaign(customer_data: dict, offer_data: dict) -> bool:
    """
    Determines if a customer and their offer are eligible for a campaign.
    Checks for DND status and active offer status.

    Args:
        customer_data (dict): Dictionary containing customer information.
                              Expected to have 'is_dnd' key.
        offer_data (dict): Dictionary containing offer information.
                           Expected to have 'offer_status' key and optionally 'valid_until'.

    Returns:
        bool: True if eligible, False otherwise.
    """
    # FR24: Avoid sending offers to DND (Do Not Disturb) customers.
    if customer_data.get('is_dnd', False):
        return False

    # FR16: Maintain flags for Offer statuses: Active, Inactive, and Expired based on defined business logic.
    # Only 'Active' offers should be eligible for campaigning.
    if offer_data.get('offer_status') != 'Active':
        return False

    # Check offer validity if 'valid_until' is provided
    if 'valid_until' in offer_data and offer_data['valid_until']:
        valid_until_dt = parse_iso_timestamp(offer_data['valid_until'])
        if valid_until_dt and valid_until_dt < datetime.datetime.now(datetime.timezone.utc):
            return False # Offer has expired

    return True