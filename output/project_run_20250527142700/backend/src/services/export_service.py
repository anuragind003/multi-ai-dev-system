import csv
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

from sqlalchemy import func

# Import db instance and models from the main application context.
# This assumes `db` is initialized in `backend/__init__.py` or `backend/app.py`
# and models are defined in `backend/models.py` based on the provided schema.
from backend import db
from backend.models import Customer, Offer, DataError


class ExportService:
    def __init__(self, db_session):
        """
        Initializes the ExportService with a SQLAlchemy database session.
        """
        self.db_session = db_session

    def generate_moengage_campaign_file(self):
        """
        Generates a Moengage-formatted CSV file for eligible customers,
        excluding DND customers and including only the latest active, non-duplicate offer.

        Functional Requirements Addressed:
        - FR30: Provide a screen for users to download the Moengage-formatted file (.csv).
        - FR24: Avoid sending offers to DND (Do Not Disturb) customers.
        - FR15: Maintain different customer attributes and customer segments.
        - FR16: Maintain flags for Offer statuses: Active, Inactive, and Expired.
        - FR17: Maintain flags for Offer types: Fresh, Enrich, New-old, New-new.
        - FR19: Maintain analytics-defined flags for Propensity.
        - FR22: Apply attribution logic (simplified here by picking latest active offer).
        - FR35: Include campaign metrics such as campaign unique identifier (placeholder).
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Define CSV headers for Moengage. These are inferred based on typical campaign
        # platform requirements and available data in the schema.
        headers = [
            "customer_id", "mobile_number", "pan_number", "offer_id", "offer_type",
            "offer_status", "valid_until", "propensity", "segment", "campaign_unique_identifier",
            "loan_application_number"
        ]
        writer.writerow(headers)

        # Subquery to find the latest updated active, non-duplicate offer for each customer.
        # This serves as a simple attribution logic (FR22) for the export, ensuring
        # only one primary offer is selected per customer for campaigning.
        latest_active_offer_subquery = self.db_session.query(
            Offer.customer_id,
            func.max(Offer.updated_at).label('latest_updated_at')
        ).filter(
            Offer.offer_status == 'Active',
            Offer.is_duplicate == False  # Exclude offers marked as duplicates
        ).group_by(Offer.customer_id).subquery()

        # Main query to join Customer and Offer tables, filtering for non-DND customers (FR24)
        # and selecting the specific latest active offer identified by the subquery.
        query = self.db_session.query(
            Customer.customer_id,
            Customer.mobile_number,
            Customer.pan_number,
            Customer.segment,
            Offer.offer_id,
            Offer.offer_type,
            Offer.offer_status,
            Offer.valid_until,
            Offer.propensity,
            Offer.loan_application_number
        ).join(
            Offer, Customer.customer_id == Offer.customer_id
        ).join(
            latest_active_offer_subquery,
            (Offer.customer_id == latest_active_offer_subquery.c.customer_id) &
            (Offer.updated_at == latest_active_offer_subquery.c.latest_updated_at)
        ).filter(
            Customer.is_dnd == False,  # Exclude DND customers (FR24)
            Offer.offer_status == 'Active',  # Only active offers (FR16)
            Offer.is_duplicate == False  # Ensure the selected offer is not a duplicate
        ).order_by(Customer.customer_id)  # Order for consistent output

        results = query.all()

        for row in results:
            # Placeholder for campaign_unique_identifier (FR35).
            # In a full implementation, this might be linked to a specific campaign
            # or generated based on campaign rules.
            campaign_unique_identifier = f"CDP_CAMPAIGN_{str(row.offer_id)[:8].upper()}"

            writer.writerow([
                str(row.customer_id),
                row.mobile_number,
                row.pan_number,
                str(row.offer_id),
                row.offer_type,
                row.offer_status,
                row.valid_until.isoformat() if row.valid_until else '',
                row.propensity,
                row.segment,
                campaign_unique_identifier,
                row.loan_application_number if row.loan_application_number else ''
            ])

        output.seek(0)
        return output.getvalue()

    def generate_duplicate_customers_file(self):
        """
        Generates a CSV file containing identified duplicate customer data.
        This includes customers associated with offers marked as 'is_duplicate=TRUE'.

        Functional Requirements Addressed:
        - FR31: Provide a screen for users to download a Duplicate Data File.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number",
            "offer_id", "offer_type", "offer_status", "is_duplicate", "original_offer_id",
            "duplicate_flag_reason"
        ]
        writer.writerow(headers)

        # Query for offers explicitly marked as duplicate and their associated customer data.
        # This implies that the deduplication service has already flagged these offers (FR3).
        query = self.db_session.query(
            Customer.customer_id,
            Customer.mobile_number,
            Customer.pan_number,
            Customer.aadhaar_number,
            Customer.ucid_number,
            Offer.offer_id,
            Offer.offer_type,
            Offer.offer_status,
            Offer.is_duplicate,
            Offer.original_offer_id
        ).join(
            Offer, Customer.customer_id == Offer.customer_id
        ).filter(
            Offer.is_duplicate == True  # Filter for offers flagged as duplicates
        ).order_by(Customer.customer_id, Offer.created_at)  # Order for consistent output

        results = query.all()

        for row in results:
            # Placeholder for duplicate_flag_reason. In a real system, this might be a column
            # in the Offer table or derived from the deduplication process.
            duplicate_flag_reason = "Marked as duplicate by CDP's deduplication logic."
            writer.writerow([
                str(row.customer_id),
                row.mobile_number,
                row.pan_number,
                row.aadhaar_number if row.aadhaar_number else '',
                row.ucid_number if row.ucid_number else '',
                str(row.offer_id),
                row.offer_type,
                row.offer_status,
                "TRUE",  # Explicitly "TRUE" as we filtered for it
                str(row.original_offer_id) if row.original_offer_id else '',
                duplicate_flag_reason
            ])

        output.seek(0)
        return output.getvalue()

    def generate_unique_customers_file(self):
        """
        Generates a CSV file containing unique customer data after deduplication.
        This file represents the single customer view (FR1).

        Functional Requirements Addressed:
        - FR32: Provide a screen for users to download a Unique Data File.
        - FR1: Perform customer deduplication to create a single profile view.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number",
            "segment", "is_dnd", "active_non_duplicate_offers_count", "latest_offer_status"
        ]
        writer.writerow(headers)

        # Query for all unique customers from the Customer table.
        # For each customer, count their active, non-duplicate offers
        # and find the status of their most recently updated offer (any type).
        query = self.db_session.query(
            Customer.customer_id,
            Customer.mobile_number,
            Customer.pan_number,
            Customer.aadhaar_number,
            Customer.ucid_number,
            Customer.segment,
            Customer.is_dnd,
            # Count active, non-duplicate offers for this customer
            func.count(Offer.offer_id).filter(
                Offer.offer_status == 'Active',
                Offer.is_duplicate == False
            ).label('active_non_duplicate_offers_count'),
            # Subquery to get the status of the single latest offer for this customer.
            # This ensures we get the actual status of the most recent offer.
            (
                self.db_session.query(Offer.offer_status)
                .filter(Offer.customer_id == Customer.customer_id)
                .order_by(Offer.updated_at.desc())
                .limit(1)
                .scalar_subquery()
            ).label('latest_offer_status')
        ).outerjoin(
            Offer, Customer.customer_id == Offer.customer_id
        ).group_by(
            Customer.customer_id,
            Customer.mobile_number,
            Customer.pan_number,
            Customer.aadhaar_number,
            Customer.ucid_number,
            Customer.segment,
            Customer.is_dnd
        ).order_by(Customer.customer_id)

        results = query.all()

        for row in results:
            writer.writerow([
                str(row.customer_id),
                row.mobile_number,
                row.pan_number,
                row.aadhaar_number if row.aadhaar_number else '',
                row.ucid_number if row.ucid_number else '',
                row.segment,
                "TRUE" if row.is_dnd else "FALSE",
                row.active_non_duplicate_offers_count,
                row.latest_offer_status if row.latest_offer_status else 'N/A'  # 'N/A' if customer has no offers
            ])

        output.seek(0)
        return output.getvalue()

    def generate_data_errors_file(self):
        """
        Generates an Excel file detailing data validation errors from ingestion processes.

        Functional Requirements Addressed:
        - FR33: Provide a screen for users to download an Error Excel file for data uploads.
        - FR2: Validate basic column-level data when moving data from Offermart to CDP.
        - NFR8: Perform basic column-level data validation during data ingestion.
        """
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data Errors"

        headers = [
            "error_id", "timestamp", "source_system", "record_identifier",
            "error_type", "error_message", "raw_data_payload"
        ]
        sheet.append(headers)

        # Apply header styling for better readability in the Excel file.
        for col_num, header_text in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Query for DataError records, ordered by timestamp descending (most recent first).
        # This assumes a DataError model exists to store ingestion errors.
        query = self.db_session.query(DataError).order_by(DataError.timestamp.desc())
        results = query.all()

        for error in results:
            row_data = [
                str(error.error_id),
                error.timestamp.isoformat() if error.timestamp else '',
                error.source_system if error.source_system else '',
                error.record_identifier if error.record_identifier else '',
                error.error_type if error.error_type else '',
                error.error_message if error.error_message else '',
                str(error.raw_data_payload) if error.raw_data_payload else ''  # Convert JSONB to string for Excel
            ]
            sheet.append(row_data)

        workbook.save(output)
        output.seek(0)
        return output.getvalue()