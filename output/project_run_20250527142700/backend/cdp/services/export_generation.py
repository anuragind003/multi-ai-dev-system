import csv
import io
from datetime import datetime
import uuid # Required for uuid.uuid4 in placeholder models
import openpyxl
from openpyxl.styles import Font, PatternFill

# Attempt to import 'db' and models from the main application structure.
# In a real Flask application, 'db' would be initialized in `backend/__init__.py`
# and models would be defined in `backend/models.py`.
try:
    from backend.__init__ import db
    from backend.models import Customer, Offer, DataError
except ImportError:
    # This block provides a fallback for isolated testing or if the full Flask app
    # context is not yet available. In a deployed application, the 'try' block
    # should succeed.
    from flask_sqlalchemy import SQLAlchemy
    from flask import Flask
    import os

    # Create a dummy Flask app and db for standalone execution/testing
    app = Flask(__name__)
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
        'DATABASE_URL',
        'postgresql://cdp_user:cdp_password@localhost:5432/cdp_db'
    )
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    db = SQLAlchemy(app)
    print("WARNING: Using mock db setup and placeholder models. Ensure 'db' is properly imported from backend.__init__ and models from backend.models in a real application.")

    # Define placeholder models here if import fails
    from sqlalchemy.dialects.postgresql import UUID as PG_UUID, JSONB
    from sqlalchemy import Column, String, Boolean, DateTime, ForeignKey, Integer, Numeric, Text, func
    from sqlalchemy.orm import relationship, declarative_base

    Base = declarative_base()

    class Customer(Base):
        __tablename__ = 'customers'
        customer_id = Column(PG_UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
        mobile_number = Column(String(20), unique=True)
        pan_number = Column(String(10), unique=True)
        aadhaar_number = Column(String(12), unique=True)
        ucid_number = Column(String(50), unique=True)
        customer_360_id = Column(String(50))
        is_dnd = Column(Boolean, default=False)
        segment = Column(String(50))
        attributes = Column(JSONB)
        created_at = Column(DateTime(timezone=True), default=func.now())
        updated_at = Column(DateTime(timezone=True), default=func.now(), onupdate=func.now())

        offers = relationship("Offer", back_populates="customer")

    class Offer(Base):
        __tablename__ = 'offers'
        offer_id = Column(PG_UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
        customer_id = Column(PG_UUID(as_uuid=True), ForeignKey('customers.customer_id'), nullable=False)
        source_offer_id = Column(String(100))
        offer_type = Column(String(50))
        offer_status = Column(String(50)) # 'Active', 'Inactive', 'Expired'
        propensity = Column(String(50))
        loan_application_number = Column(String(100))
        valid_until = Column(DateTime(timezone=True))
        source_system = Column(String(50))
        channel = Column(String(50))
        is_duplicate = Column(Boolean, default=False)
        original_offer_id = Column(PG_UUID(as_uuid=True), ForeignKey('offers.offer_id')) # Self-referencing
        created_at = Column(DateTime(timezone=True), default=func.now())
        updated_at = Column(DateTime(timezone=True), default=func.now(), onupdate=func.now())

        customer = relationship("Customer", back_populates="offers")
        original_offer = relationship("Offer", remote_side=[offer_id], uselist=False) # For original_offer_id

    class DataError(Base):
        __tablename__ = 'data_errors'
        error_id = Column(PG_UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
        source_file_name = Column(String(255))
        row_number = Column(Integer)
        column_name = Column(String(255))
        error_message = Column(Text)
        original_value = Column(Text)
        error_timestamp = Column(DateTime(timezone=True), default=func.now())


class ExportGenerationService:
    """
    Service class responsible for generating various data export files
    (Moengage campaign file, duplicate customer data, unique customer data, data errors).
    """

    @staticmethod
    def generate_moengage_campaign_file():
        """
        Generates a Moengage-formatted CSV file for eligible customers.
        Excludes DND customers and includes only active offers.

        Functional Requirements Addressed:
        - FR30: The CDP system shall provide a screen for users to download the Moengage-formatted file (.csv).
        - FR24: The CDP system shall avoid sending offers to DND (Do Not Disturb) customers.
        - FR16: The CDP system shall maintain flags for Offer statuses: Active, Inactive, and Expired based on defined business logic.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Moengage specific headers (example, adjust as per actual Moengage requirements)
        headers = [
            "customer_id", "mobile_number", "pan_number", "offer_id",
            "offer_type", "offer_status", "propensity", "loan_application_number",
            "valid_until", "segment", "source_system", "channel"
        ]
        writer.writerow(headers)

        try:
            # Query for active offers for non-DND customers
            eligible_offers = db.session.query(Customer, Offer)\
                .join(Offer, Customer.customer_id == Offer.customer_id)\
                .filter(Customer.is_dnd == False)\
                .filter(Offer.offer_status == 'Active')\
                .order_by(Customer.customer_id, Offer.created_at.desc())\
                .all()

            for customer, offer in eligible_offers:
                row = [
                    str(customer.customer_id),
                    customer.mobile_number,
                    customer.pan_number,
                    str(offer.offer_id),
                    offer.offer_type,
                    offer.offer_status,
                    offer.propensity,
                    offer.loan_application_number,
                    offer.valid_until.isoformat() if offer.valid_until else '',
                    customer.segment,
                    offer.source_system,
                    offer.channel
                ]
                writer.writerow(row)

            output.seek(0)
            return output.getvalue()
        except Exception as e:
            # In a production system, use a proper logging framework
            print(f"Error generating Moengage campaign file: {e}")
            raise

    @staticmethod
    def generate_duplicate_customers_file():
        """
        Generates a CSV file containing identified duplicate customer data.
        This typically means offers/customer records that were flagged during deduplication.

        Functional Requirements Addressed:
        - FR31: The CDP system shall provide a screen for users to download a Duplicate Data File.
        - FR1: The CDP system shall perform customer deduplication to create a single profile view for Consumer Loan Products.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number",
            "offer_id", "source_offer_id", "offer_type", "offer_status", "is_duplicate",
            "original_offer_id", "duplicate_flag_date"
        ]
        writer.writerow(headers)

        try:
            # Query for offers explicitly marked as duplicate
            duplicate_offers = db.session.query(Customer, Offer)\
                .join(Offer, Customer.customer_id == Offer.customer_id)\
                .filter(Offer.is_duplicate == True)\
                .order_by(Customer.customer_id, Offer.created_at)\
                .all()

            for customer, offer in duplicate_offers:
                row = [
                    str(customer.customer_id),
                    customer.mobile_number,
                    customer.pan_number,
                    customer.aadhaar_number,
                    customer.ucid_number,
                    str(offer.offer_id),
                    offer.source_offer_id,
                    offer.offer_type,
                    offer.offer_status,
                    offer.is_duplicate,
                    str(offer.original_offer_id) if offer.original_offer_id else '',
                    offer.updated_at.isoformat() if offer.updated_at else '' # Using updated_at as proxy for flag date
                ]
                writer.writerow(row)

            output.seek(0)
            return output.getvalue()
        except Exception as e:
            print(f"Error generating duplicate customers file: {e}")
            raise

    @staticmethod
    def generate_unique_customers_file():
        """
        Generates a CSV file containing unique customer data after deduplication.
        This typically means the primary customer profiles and their associated active offers.

        Functional Requirements Addressed:
        - FR32: The CDP system shall provide a screen for users to download a Unique Data File.
        - FR1: The CDP system shall perform customer deduplication to create a single profile view for Consumer Loan Products.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number",
            "segment", "is_dnd", "active_offers_count", "last_offer_update"
        ]
        writer.writerow(headers)

        try:
            # Query for unique customers. Since `customers` table itself represents unique profiles
            # after deduplication, we query it directly. We aggregate some offer info for context.
            unique_customers_query = db.session.query(
                Customer,
                func.count(Offer.offer_id).label('active_offers_count'),
                func.max(Offer.updated_at).label('last_offer_update')
            ).outerjoin(Offer, (Customer.customer_id == Offer.customer_id) & (Offer.offer_status == 'Active'))\
            .group_by(Customer.customer_id)\
            .order_by(Customer.created_at)\
            .all()

            for customer, active_offers_count, last_offer_update in unique_customers_query:
                row = [
                    str(customer.customer_id),
                    customer.mobile_number,
                    customer.pan_number,
                    customer.aadhaar_number,
                    customer.ucid_number,
                    customer.segment,
                    customer.is_dnd,
                    active_offers_count,
                    last_offer_update.isoformat() if last_offer_update else ''
                ]
                writer.writerow(row)

            output.seek(0)
            return output.getvalue()
        except Exception as e:
            print(f"Error generating unique customers file: {e}")
            raise

    @staticmethod
    def generate_data_errors_file():
        """
        Generates an Excel file detailing data validation errors from ingestion processes.

        Functional Requirements Addressed:
        - FR33: The CDP system shall provide a screen for users to download an Error Excel file for data uploads.
        """
        # Create a new workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data Errors"

        headers = [
            "Error ID", "Source File Name", "Row Number", "Column Name",
            "Error Message", "Original Value", "Error Timestamp"
        ]
        sheet.append(headers)

        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        try:
            # Query all data errors
            errors = db.session.query(DataError).order_by(DataError.error_timestamp.desc()).all()

            for error in errors:
                row = [
                    str(error.error_id),
                    error.source_file_name,
                    error.row_number,
                    error.column_name,
                    error.error_message,
                    error.original_value,
                    error.error_timestamp.isoformat() if error.error_timestamp else ''
                ]
                sheet.append(row)

            # Auto-size columns for better readability
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name (e.g., 'A', 'B')
                for cell in col:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except TypeError: # Handle cases where cell.value might not be string-convertible
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 0: # Ensure width is positive
                    sheet.column_dimensions[column].width = adjusted_width

            # Save the workbook to a BytesIO object
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            print(f"Error generating data errors file: {e}")
            raise