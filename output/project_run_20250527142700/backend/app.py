import os
import uuid
import datetime
import io
import csv
from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from sqlalchemy.dialects.postgresql import UUID, JSONB
from sqlalchemy import text, func, or_
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

# --- Flask App Setup ---
app = Flask(__name__)
CORS(app) # Enable CORS for frontend integration

# Database Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://cdp_user:cdp_password@localhost:5432/cdp_db'
).replace('postgres://', 'postgresql://') # SQLAlchemy 1.4+ requires postgresql://

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your_super_secret_key_here') # Replace with a strong secret key in production

db = SQLAlchemy(app)

# --- Database Models ---

class Customer(db.Model):
    __tablename__ = 'customers'
    customer_id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    mobile_number = db.Column(db.String(20), unique=True, nullable=True)
    pan_number = db.Column(db.String(10), unique=True, nullable=True)
    aadhaar_number = db.Column(db.String(12), unique=True, nullable=True)
    ucid_number = db.Column(db.String(50), unique=True, nullable=True)
    customer_360_id = db.Column(db.String(50), nullable=True) # For integration with Customer 360
    is_dnd = db.Column(db.Boolean, default=False)
    segment = db.Column(db.String(50), nullable=True) # C1-C8, etc.
    attributes = db.Column(JSONB, nullable=True) # For other customer attributes
    created_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc))
    updated_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc), onupdate=datetime.datetime.now(datetime.timezone.utc))

    offers = db.relationship('Offer', backref='customer', lazy=True)
    events = db.relationship('Event', backref='customer', lazy=True)

    def __repr__(self):
        return f"<Customer {self.customer_id} - {self.mobile_number}>"

class Offer(db.Model):
    __tablename__ = 'offers'
    offer_id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    customer_id = db.Column(UUID(as_uuid=True), db.ForeignKey('customers.customer_id'), nullable=False)
    source_offer_id = db.Column(db.String(100), nullable=True) # Original ID from Offermart/E-aggregator
    offer_type = db.Column(db.String(50), nullable=True) # 'Fresh', 'Enrich', 'New-old', 'New-new'
    offer_status = db.Column(db.String(50), default='Active') # 'Active', 'Inactive', 'Expired'
    propensity = db.Column(db.String(50), nullable=True)
    loan_application_number = db.Column(db.String(100), nullable=True) # LAN
    valid_until = db.Column(db.DateTime(timezone=True), nullable=True)
    source_system = db.Column(db.String(50), nullable=False) # 'Offermart', 'E-aggregator'
    channel = db.Column(db.String(50), nullable=True) # For attribution
    is_duplicate = db.Column(db.Boolean, default=False) # Flagged by deduplication
    original_offer_id = db.Column(UUID(as_uuid=True), db.ForeignKey('offers.offer_id'), nullable=True) # Points to the offer it duplicated/enriched
    created_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc))
    updated_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc), onupdate=datetime.datetime.now(datetime.timezone.utc))

    history = db.relationship('OfferHistory', backref='offer', lazy=True)
    events = db.relationship('Event', backref='offer', lazy=True)

    def __repr__(self):
        return f"<Offer {self.offer_id} for Customer {self.customer_id} - {self.offer_status}>"

class OfferHistory(db.Model):
    __tablename__ = 'offer_history'
    history_id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    offer_id = db.Column(UUID(as_uuid=True), db.ForeignKey('offers.offer_id'), nullable=False)
    status_change_date = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc))
    old_status = db.Column(db.String(50), nullable=True)
    new_status = db.Column(db.String(50), nullable=False)
    change_reason = db.Column(db.Text, nullable=True)

    def __repr__(self):
        return f"<OfferHistory {self.history_id} - Offer {self.offer_id} changed to {self.new_status}>"

class Event(db.Model):
    __tablename__ = 'events'
    event_id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    customer_id = db.Column(UUID(as_uuid=True), db.ForeignKey('customers.customer_id'), nullable=True)
    offer_id = db.Column(UUID(as_uuid=True), db.ForeignKey('offers.offer_id'), nullable=True)
    event_type = db.Column(db.String(100), nullable=False) # SMS_SENT, EKYC_ACHIEVED, JOURNEY_LOGIN, etc.
    event_timestamp = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc))
    source_system = db.Column(db.String(50), nullable=False) # Moengage, LOS
    event_details = db.Column(JSONB, nullable=True) # Raw event payload

    def __repr__(self):
        return f"<Event {self.event_id} - {self.event_type} from {self.source_system}>"

class Campaign(db.Model):
    __tablename__ = 'campaigns'
    campaign_id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    campaign_name = db.Column(db.String(255), nullable=False)
    campaign_date = db.Column(db.Date, nullable=False)
    campaign_unique_identifier = db.Column(db.String(100), unique=True, nullable=False)
    attempted_count = db.Column(db.Integer, default=0)
    sent_count = db.Column(db.Integer, default=0)
    failed_count = db.Column(db.Integer, default=0)
    success_rate = db.Column(db.Numeric(5,2), default=0.0)
    conversion_rate = db.Column(db.Numeric(5,2), default=0.0)
    created_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc))
    updated_at = db.Column(db.DateTime(timezone=True), default=datetime.datetime.now(datetime.timezone.utc), onupdate=datetime.datetime.now(datetime.timezone.utc))

    def __repr__(self):
        return f"<Campaign {self.campaign_id} - {self.campaign_name}>"

# --- Helper Functions / Business Logic ---

def validate_data(data, required_fields):
    """Performs basic column-level data validation."""
    errors = []
    for field in required_fields:
        if field not in data or not data[field]:
            errors.append(f"Missing or empty required field: {field}")
    return errors

def deduplicate_customer(customer_data):
    """
    Deduplicates customers based on Mobile, Pan, Aadhaar, UCID. (FR1, FR3, FR4, FR5, FR6)
    Returns existing customer object if found, otherwise None.
    """
    mobile = customer_data.get('mobile_number')
    pan = customer_data.get('pan_number')
    aadhaar = customer_data.get('aadhaar_number')
    ucid = customer_data.get('ucid_number')

    query_filters = []
    if mobile:
        query_filters.append(Customer.mobile_number == mobile)
    if pan:
        query_filters.append(Customer.pan_number == pan)
    if aadhaar:
        query_filters.append(Customer.aadhaar_number == aadhaar)
    if ucid:
        query_filters.append(Customer.ucid_number == ucid)

    if not query_filters:
        return None # No identifiable data for deduplication

    existing_customer = Customer.query.filter(or_(*query_filters)).first()
    return existing_customer

def update_offer_status(offer_id, new_status, reason=""):
    """Updates an offer's status and logs it in offer_history."""
    offer = Offer.query.get(offer_id)
    if not offer:
        return False

    old_status = offer.offer_status
    offer.offer_status = new_status
    db.session.add(offer)

    history_entry = OfferHistory(
        offer_id=offer.offer_id,
        old_status=old_status,
        new_status=new_status,
        change_reason=reason
    )
    db.session.add(history_entry)
    # Commit is handled by the calling function's transaction
    return True

def apply_attribution_logic(customer_id, new_offer_data):
    """
    Applies attribution logic (FR22).
    Simplified: For MVP, if a customer has multiple active offers, the newest 'Fresh' offer prevails,
    or the one from a preferred channel (e.g., E-aggregator over Offermart).
    This is a placeholder for more complex business rules.
    """
    # Example: Mark older active offers as 'Inactive' if a new 'Fresh' offer comes in
    # This logic needs detailed business rules. For now, we'll just return True,
    # implying the new offer is considered.
    # In a real scenario, this would involve comparing offer types, channels, values, etc.
    return True

# --- Data Export Functions ---

def generate_moengage_csv():
    """Generates a Moengage-formatted CSV file for eligible customers (FR30, FR24)."""
    si = io.StringIO()
    cw = csv.writer(si)

    # Headers for Moengage file (example, actual headers need to be defined)
    headers = [
        "customer_id", "mobile_number", "pan_number", "offer_id", "offer_type",
        "offer_status", "valid_until", "propensity", "segment", "campaign_id"
    ]
    cw.writerow(headers)

    # Fetch active offers for non-DND customers
    eligible_offers = db.session.query(Offer, Customer).join(Customer).filter(
        Offer.offer_status == 'Active',
        Customer.is_dnd == False
    ).all()

    for offer, customer in eligible_offers:
        # Placeholder for campaign_id - in a real system, this would come from a campaign run
        campaign_id = "CDP_CAMPAIGN_DEFAULT"
        row = [
            str(customer.customer_id),
            customer.mobile_number,
            customer.pan_number,
            str(offer.offer_id),
            offer.offer_type,
            offer.offer_status,
            offer.valid_until.isoformat() if offer.valid_until else '',
            offer.propensity,
            customer.segment,
            campaign_id
        ]
        cw.writerow(row)

    output = io.BytesIO(si.getvalue().encode('utf-8'))
    return output

def generate_duplicate_csv():
    """Generates a CSV file for identified duplicate customer data (FR31)."""
    si = io.StringIO()
    cw = csv.writer(si)

    headers = ["customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number", "duplicate_of_customer_id", "offer_id", "source_offer_id"]
    cw.writerow(headers)

    # Fetch offers that were marked as duplicate
    duplicate_offers = Offer.query.filter_by(is_duplicate=True).all()

    for offer in duplicate_offers:
        customer = Customer.query.get(offer.customer_id)
        original_offer_customer_id = None
        if offer.original_offer_id:
            original_offer = Offer.query.get(offer.original_offer_id)
            if original_offer:
                original_offer_customer_id = str(original_offer.customer_id)

        row = [
            str(customer.customer_id) if customer else '',
            customer.mobile_number if customer else '',
            customer.pan_number if customer else '',
            customer.aadhaar_number if customer else '',
            customer.ucid_number if customer else '',
            original_offer_customer_id,
            str(offer.offer_id),
            offer.source_offer_id
        ]
        cw.writerow(row)

    output = io.BytesIO(si.getvalue().encode('utf-8'))
    return output

def generate_unique_csv():
    """Generates a CSV file for unique customer data after deduplication (FR32)."""
    si = io.StringIO()
    cw = csv.writer(si)

    headers = ["customer_id", "mobile_number", "pan_number", "aadhaar_number", "ucid_number", "segment", "active_offers_count"]
    cw.writerow(headers)

    # Fetch customers who are not marked as duplicates (or whose offers are not duplicates)
    # For simplicity, we'll consider customers with at least one non-duplicate active offer.
    unique_customers = db.session.query(Customer).join(Offer).filter(
        Offer.is_duplicate == False,
        Offer.offer_status == 'Active'
    ).distinct().all()

    for customer in unique_customers:
        active_offers_count = Offer.query.filter_by(customer_id=customer.customer_id, offer_status='Active').count()
        row = [
            str(customer.customer_id),
            customer.mobile_number,
            customer.pan_number,
            customer.aadhaar_number,
            customer.ucid_number,
            customer.segment,
            active_offers_count
        ]
        cw.writerow(row)

    output = io.BytesIO(si.getvalue().encode('utf-8'))
    return output

# In a real system, errors would be stored in a database table.
# For this MVP, we'll simulate a list of errors.
_data_errors_log = []

def log_data_error(source, record_id, error_message, payload):
    """Logs data validation errors."""
    _data_errors_log.append({
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "source": source,
        "record_id": record_id,
        "error_message": error_message,
        "payload": payload
    })

def generate_error_excel():
    """Generates an Excel file detailing data validation errors (FR33)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Errors"

    headers = ["Timestamp", "Source System", "Record ID", "Error Message", "Payload"]
    ws.append(headers)

    # Apply header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for error in _data_errors_log:
        ws.append([
            error.get("timestamp"),
            error.get("source"),
            error.get("record_id"),
            error.get("error_message"),
            str(error.get("payload")) # Convert dict to string for Excel cell
        ])

    # Auto-size columns (basic attempt)
    for col_idx, col in enumerate(ws.columns):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- API Endpoints ---

@app.route('/ingest/e-aggregator-data', methods=['POST'])
def ingest_e_aggregator_data():
    """
    Receives real-time lead, eligibility, or status data from E-aggregators.
    Performs basic validation and inserts/updates customer and offer data. (FR10, FR11)
    """
    data = request.get_json()
    if not data:
        log_data_error("E-Aggregator", "N/A", "No JSON payload provided", request.data.decode())
        return jsonify({"status": "error", "message": "Invalid JSON payload"}), 400

    required_fields = ['source_system', 'data_type', 'payload']
    errors = validate_data(data, required_fields)
    if errors:
        log_data_error("E-Aggregator", data.get('payload', {}).get('id', 'N/A'), f"Validation errors: {', '.join(errors)}", data)
        return jsonify({"status": "error", "message": "Validation failed", "errors": errors}), 400

    source_system = data['source_system']
    data_type = data['data_type']
    payload = data['payload']

    # Extract relevant customer and offer data from payload
    # This mapping will be highly dependent on the actual E-aggregator payload structure
    customer_data = {
        'mobile_number': payload.get('mobile_number'),
        'pan_number': payload.get('pan_number'),
        'aadhaar_number': payload.get('aadhaar_number'),
        'ucid_number': payload.get('ucid_number'),
        'segment': payload.get('segment'),
        'attributes': payload.get('customer_attributes')
    }
    offer_data = {
        'source_offer_id': payload.get('offer_id'),
        'offer_type': payload.get('offer_type', 'Fresh'), # Default to 'Fresh' if not specified
        'propensity': payload.get('propensity'),
        'loan_application_number': payload.get('loan_application_number'),
        'valid_until': datetime.datetime.fromisoformat(payload['valid_until']) if 'valid_until' in payload else None,
        'source_system': source_system,
        'channel': payload.get('channel', source_system)
    }

    try:
        # 1. Deduplicate Customer (FR1, FR3, FR4, FR5)
        existing_customer = deduplicate_customer(customer_data)
        customer_obj = None
        if existing_customer:
            customer_obj = existing_customer
            # Update existing customer attributes if necessary
            if customer_data.get('segment'):
                customer_obj.segment = customer_data['segment']
            if customer_data.get('attributes'):
                customer_obj.attributes = {**(customer_obj.attributes or {}), **customer_data['attributes']}
            db.session.add(customer_obj)
        else:
            # Create new customer
            customer_obj = Customer(**{k: v for k, v in customer_data.items() if v is not None})
            db.session.add(customer_obj)
        db.session.flush() # To get customer_id for new customer

        # 2. Handle Offer Logic
        # Check for active loan application journey (FR13)
        if offer_data.get('loan_application_number'):
            existing_offer_with_lan = Offer.query.filter_by(
                customer_id=customer_obj.customer_id,
                loan_application_number=offer_data['loan_application_number'],
                offer_status='Active' # Assuming 'Active' status implies journey started
            ).first()
            if existing_offer_with_lan:
                log_data_error(source_system, offer_data['source_offer_id'],
                               "Offer modification prevented: active loan application journey.", payload)
                return jsonify({"status": "error", "message": "Offer modification prevented due to active loan application journey"}), 409

        # Check for 'Enrich' offers (FR18)
        original_offer_id_for_enrich = None
        if offer_data.get('offer_type') == 'Enrich':
            # Find previous active offer for this customer
            previous_active_offer = Offer.query.filter_by(
                customer_id=customer_obj.customer_id,
                offer_status='Active'
            ).order_by(Offer.created_at.desc()).first()

            if previous_active_offer:
                # If journey not started for previous offer, mark previous as Inactive/Duplicate
                # (Simplified check: if no LAN or LAN is not active, assume journey not started)
                if not previous_active_offer.loan_application_number or \
                   (previous_active_offer.loan_application_number and
                    not Event.query.filter_by(
                        loan_application_number=previous_active_offer.loan_application_number,
                        event_type='JOURNEY_LOGIN' # Or other journey start event
                    ).first()):
                    update_offer_status(previous_active_offer.offer_id, 'Inactive', 'Enriched by new offer')
                    offer_data['is_duplicate'] = True
                    original_offer_id_for_enrich = previous_active_offer.offer_id
                else:
                    # If journey started, do not flow to CDP (or reject this 'Enrich' offer)
                    log_data_error(source_system, offer_data['source_offer_id'],
                                   "Enrich offer rejected: previous offer has active journey.", payload)
                    return jsonify({"status": "error", "message": "Enrich offer rejected due to active journey on previous offer"}), 409

        # Apply attribution logic (FR22) - this might involve marking other offers inactive
        apply_attribution_logic(customer_obj.customer_id, offer_data)

        # Create new offer
        new_offer = Offer(
            customer_id=customer_obj.customer_id,
            original_offer_id=original_offer_id_for_enrich,
            **{k: v for k, v in offer_data.items() if v is not None}
        )
        db.session.add(new_offer)

        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Data processed successfully",
            "customer_id": str(customer_obj.customer_id),
            "offer_id": str(new_offer.offer_id)
        }), 200

    except Exception as e:
        db.session.rollback()
        log_data_error(source_system, payload.get('id', 'N/A'), f"Internal server error: {str(e)}", payload)
        app.logger.error(f"Error processing E-aggregator data: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Internal server error", "details": str(e)}), 500

@app.route('/events/moengage', methods=['POST'])
def receive_moengage_event():
    """Receives SMS campaign events (sent, delivered, click) from Moengage (FR23, FR25)."""
    data = request.get_json()
    if not data:
        log_data_error("Moengage", "N/A", "No JSON payload provided", request.data.decode())
        return jsonify({"status": "error", "message": "Invalid JSON payload"}), 400

    required_fields = ['customer_mobile', 'event_type', 'timestamp', 'campaign_id']
    errors = validate_data(data, required_fields)
    if errors:
        log_data_error("Moengage", data.get('campaign_id', 'N/A'), f"Validation errors: {', '.join(errors)}", data)
        return jsonify({"status": "error", "message": "Validation failed", "errors": errors}), 400

    try:
        customer = Customer.query.filter_by(mobile_number=data['customer_mobile']).first()
        customer_id = customer.customer_id if customer else None

        # Try to link event to an offer, e.g., by campaign_id or recent active offer
        offer = None
        if customer_id:
            # This is a simplified lookup. A real system might need to match campaign_id to offer_id
            # or find the most recent offer associated with the campaign.
            offer = Offer.query.filter_by(customer_id=customer_id, offer_status='Active').order_by(Offer.created_at.desc()).first()

        event = Event(
            customer_id=customer_id,
            offer_id=offer.offer_id if offer else None,
            event_type=data['event_type'],
            event_timestamp=datetime.datetime.fromisoformat(data['timestamp']),
            source_system='Moengage',
            event_details=data.get('details', {})
        )
        db.session.add(event)
        db.session.commit()

        return jsonify({"status": "success", "message": "Moengage event recorded"}), 200
    except Exception as e:
        db.session.rollback()
        log_data_error("Moengage", data.get('campaign_id', 'N/A'), f"Internal server error: {str(e)}", data)
        app.logger.error(f"Error processing Moengage event: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Internal server error", "details": str(e)}), 500

@app.route('/events/los', methods=['POST'])
def receive_los_event():
    """Receives loan application journey and conversion events from LOS (FR23, FR26, FR27)."""
    data = request.get_json()
    if not data:
        log_data_error("LOS", "N/A", "No JSON payload provided", request.data.decode())
        return jsonify({"status": "error", "message": "Invalid JSON payload"}), 400

    required_fields = ['loan_application_number', 'event_type', 'timestamp']
    errors = validate_data(data, required_fields)
    if errors:
        log_data_error("LOS", data.get('loan_application_number', 'N/A'), f"Validation errors: {', '.join(errors)}", data)
        return jsonify({"status": "error", "message": "Validation failed", "errors": errors}), 400

    try:
        loan_application_number = data['loan_application_number']
        event_type = data['event_type']
        event_timestamp = datetime.datetime.fromisoformat(data['timestamp'])

        # Find the associated offer by loan_application_number
        offer = Offer.query.filter_by(loan_application_number=loan_application_number).first()
        customer_id = offer.customer_id if offer else None

        event = Event(
            customer_id=customer_id,
            offer_id=offer.offer_id if offer else None,
            event_type=event_type,
            event_timestamp=event_timestamp,
            source_system='LOS',
            event_details=data.get('details', {})
        )
        db.session.add(event)

        # Update offer status if LAN validity is over (FR36)
        # This logic needs to be more specific about what 'LAN validity is over' means.
        # For now, if a 'Rejected' or 'Expired' event comes for a LAN, mark the offer as Expired.
        if offer and event_type in ['LOAN_REJECTED', 'LOAN_EXPIRED', 'LOAN_DISBURSED']:
            # If disbursed, offer is fulfilled, can be marked inactive/expired
            # If rejected/expired, offer is no longer valid
            update_offer_status(offer.offer_id, 'Expired', f"Loan application {event_type}")

        db.session.commit()

        return jsonify({"status": "success", "message": "LOS event recorded"}), 200
    except Exception as e:
        db.session.rollback()
        log_data_error("LOS", data.get('loan_application_number', 'N/A'), f"Internal server error: {str(e)}", data)
        app.logger.error(f"Error processing LOS event: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Internal server error", "details": str(e)}), 500

@app.route('/customers/<uuid:customer_id>', methods=['GET'])
def get_customer_profile(customer_id):
    """Retrieves a single customer's de-duplicated profile and associated active offers (FR1, FR16)."""
    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"status": "error", "message": "Customer not found"}), 404

    active_offers = Offer.query.filter_by(customer_id=customer.customer_id, offer_status='Active').all()

    customer_data = {
        "customer_id": str(customer.customer_id),
        "mobile_number": customer.mobile_number,
        "pan_number": customer.pan_number,
        "aadhaar_number": customer.aadhaar_number,
        "ucid_number": customer.ucid_number,
        "is_dnd": customer.is_dnd,
        "segment": customer.segment,
        "attributes": customer.attributes,
        "created_at": customer.created_at.isoformat(),
        "updated_at": customer.updated_at.isoformat(),
        "offers": [
            {
                "offer_id": str(offer.offer_id),
                "source_offer_id": offer.source_offer_id,
                "offer_type": offer.offer_type,
                "offer_status": offer.offer_status,
                "propensity": offer.propensity,
                "loan_application_number": offer.loan_application_number,
                "valid_until": offer.valid_until.isoformat() if offer.valid_until else None,
                "source_system": offer.source_system,
                "channel": offer.channel,
                "is_duplicate": offer.is_duplicate,
                "original_offer_id": str(offer.original_offer_id) if offer.original_offer_id else None,
                "created_at": offer.created_at.isoformat(),
                "updated_at": offer.updated_at.isoformat()
            } for offer in active_offers
        ]
    }
    return jsonify({"status": "success", "data": customer_data}), 200

@app.route('/exports/moengage-campaign-file', methods=['GET'])
def export_moengage_campaign_file():
    """Generates and allows download of a Moengage-formatted CSV file (FR30)."""
    try:
        csv_buffer = generate_moengage_csv()
        return send_file(
            csv_buffer,
            mimetype='text/csv',
            as_attachment=True,
            download_name='moengage_campaign_data.csv'
        )
    except Exception as e:
        app.logger.error(f"Error generating Moengage CSV: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to generate Moengage CSV", "details": str(e)}), 500

@app.route('/exports/duplicate-customers', methods=['GET'])
def export_duplicate_customers():
    """Generates and allows download of a file containing identified duplicate customer data (FR31)."""
    try:
        csv_buffer = generate_duplicate_csv()
        return send_file(
            csv_buffer,
            mimetype='text/csv',
            as_attachment=True,
            download_name='duplicate_customers.csv'
        )
    except Exception as e:
        app.logger.error(f"Error generating duplicate customers CSV: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to generate duplicate customers CSV", "details": str(e)}), 500

@app.route('/exports/unique-customers', methods=['GET'])
def export_unique_customers():
    """Generates and allows download of a file containing unique customer data (FR32)."""
    try:
        csv_buffer = generate_unique_csv()
        return send_file(
            csv_buffer,
            mimetype='text/csv',
            as_attachment=True,
            download_name='unique_customers.csv'
        )
    except Exception as e:
        app.logger.error(f"Error generating unique customers CSV: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to generate unique customers CSV", "details": str(e)}), 500

@app.route('/exports/data-errors', methods=['GET'])
def export_data_errors():
    """Generates and allows download of an Excel file detailing data validation errors (FR33)."""
    try:
        excel_buffer = generate_error_excel()
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='data_errors.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Error generating data errors Excel: {e}", exc_info=True)
        return jsonify({"status": "error", "message": "Failed to generate data errors Excel", "details": str(e)}), 500

# --- Database Initialization (for development/testing) ---
# These commands can be run using `flask create-db` or `flask drop-db`
@app.cli.command('create-db')
def create_db_command():
    """Creates database tables."""
    with app.app_context():
        db.create_all()
        print("Database tables created.")

@app.cli.command('drop-db')
def drop_db_command():
    """Drops database tables."""
    with app.app_context():
        db.drop_all()
        print("Database tables dropped.")

# --- Main entry point for running the app ---
if __name__ == '__main__':
    # This block is typically used for local development.
    # For production, a WSGI server like Gunicorn would be used.
    app.run(debug=True, host='0.0.0.0', port=5000)