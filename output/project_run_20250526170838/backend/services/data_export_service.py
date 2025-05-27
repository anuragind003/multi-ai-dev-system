import io
import csv
from datetime import datetime, timezone
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill

# In a typical Flask application, `db` would be initialized in `backend/extensions.py`
# and models would be defined in `backend/models.py`.
# For the purpose of providing a complete, runnable file as requested,
# `db` and the SQLAlchemy models are defined here.
# In a real project, you would import them like:
# from backend.extensions import db
# from backend.models import Customer, Offer, IngestionLog, etc.

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.dialects.postgresql import JSONB

db = SQLAlchemy() # This instance needs to be initialized with a Flask app later.

# --- SQLAlchemy Models (Mirroring the provided schema) ---
# These models would typically reside in `backend/models.py`

class Customer(db.Model):
    __tablename__ = 'customers'
    customer_id = db.Column(db.Text, primary_key=True)
    mobile_number = db.Column(db.Text, unique=True)
    pan_number = db.Column(db.Text, unique=True)
    aadhaar_number = db.Column(db.Text, unique=True)
    ucid_number = db.Column(db.Text, unique=True)
    loan_application_number = db.Column(db.Text, unique=True)
    dnd_flag = db.Column(db.Boolean, default=False)
    segment = db.Column(db.Text)
    created_at = db.Column(db.TIMESTAMP, default=datetime.now(timezone.utc))
    updated_at = db.Column(
        db.TIMESTAMP,
        default=datetime.now(timezone.utc),
        onupdate=datetime.now(timezone.utc)
    )

    offers = db.relationship('Offer', backref='customer', lazy=True)
    events = db.relationship('Event', backref='customer', lazy=True)


class Offer(db.Model):
    __tablename__ = 'offers'
    offer_id = db.Column(db.Text, primary_key=True)
    customer_id = db.Column(db.Text, db.ForeignKey('customers.customer_id'),
                            nullable=False)
    offer_type = db.Column(db.Text)
    offer_status = db.Column(db.Text)
    propensity = db.Column(db.Text)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    channel = db.Column(db.Text)
    created_at = db.Column(db.TIMESTAMP, default=datetime.now(timezone.utc))
    updated_at = db.Column(
        db.TIMESTAMP,
        default=datetime.now(timezone.utc),
        onupdate=datetime.now(timezone.utc)
    )


class Event(db.Model):
    __tablename__ = 'events'
    event_id = db.Column(db.Text, primary_key=True)
    customer_id = db.Column(db.Text, db.ForeignKey('customers.customer_id'),
                            nullable=False)
    event_type = db.Column(db.Text)
    event_source = db.Column(db.Text)
    event_timestamp = db.Column(db.TIMESTAMP)
    event_details = db.Column(JSONB)
    created_at = db.Column(db.TIMESTAMP, default=datetime.now(timezone.utc))


class CampaignMetric(db.Model):
    __tablename__ = 'campaign_metrics'
    metric_id = db.Column(db.Text, primary_key=True)
    campaign_unique_id = db.Column(db.Text, unique=True, nullable=False)
    campaign_name = db.Column(db.Text)
    campaign_date = db.Column(db.Date)
    attempted_count = db.Column(db.Integer)
    sent_success_count = db.Column(db.Integer)
    failed_count = db.Column(db.Integer)
    conversion_rate = db.Column(db.Numeric(5, 2))
    created_at = db.Column(db.TIMESTAMP, default=datetime.now(timezone.utc))


class IngestionLog(db.Model):
    __tablename__ = 'ingestion_logs'
    log_id = db.Column(db.Text, primary_key=True)
    file_name = db.Column(db.Text, nullable=False)
    upload_timestamp = db.Column(db.TIMESTAMP,
                                 default=datetime.now(timezone.utc))
    status = db.Column(db.Text)
    error_description = db.Column(db.Text)


# --- End SQLAlchemy Models ---


class DataExportService:
    """
    Service class for handling various data export functionalities.
    """

    def export_moengage_format_csv(self) -> io.StringIO:
        """
        Generates a CSV file in Moengage format, excluding DND customers.
        (FR31, FR44, FR23)
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Moengage format fields (assumed based on common campaign needs)
        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number",
            "offer_id", "offer_type", "propensity", "offer_start_date",
            "offer_end_date", "segment", "channel"
        ]
        writer.writerow(headers)

        # Query customers who are not DND and have active offers
        customers_with_offers = db.session.query(Customer, Offer).join(
            Offer, Customer.customer_id == Offer.customer_id
        ).filter(
            not Customer.dnd_flag,
            Offer.offer_status == 'Active'
        ).all()

        for customer, offer in customers_with_offers:
            row = [
                customer.customer_id,
                customer.mobile_number,
                customer.pan_number,
                customer.aadhaar_number,
                offer.offer_id,
                offer.offer_type,
                offer.propensity,
                offer.start_date.strftime('%Y-%m-%d')
                if offer.start_date else '',
                offer.end_date.strftime('%Y-%m-%d')
                if offer.end_date else '',
                customer.segment,
                offer.channel
            ]
            writer.writerow(row)

        output.seek(0)
        return output

    def export_duplicate_data_csv(self) -> io.StringIO:
        """
        Generates a CSV file containing identified duplicate customer records.
        (FR32)

        ASSUMPTION: A `duplicate_customer_records` table exists to store
        the original data of records identified as duplicates during ingestion.
        This table is not explicitly in the provided schema but is necessary
        to fulfill FR32. In a real system, this table would be populated
        by the deduplication engine.
        """
        # Define the assumed model for duplicate records
        class DuplicateCustomerRecord(db.Model):
            __tablename__ = 'duplicate_customer_records'
            original_customer_id = db.Column(db.Text, primary_key=True)
            mobile_number = db.Column(db.Text)
            pan_number = db.Column(db.Text)
            aadhaar_number = db.Column(db.Text)
            ucid_number = db.Column(db.Text)
            loan_application_number = db.Column(db.Text)
            reason_for_duplication = db.Column(db.Text)
            master_customer_id = db.Column(
                db.Text, db.ForeignKey('customers.customer_id'))
            ingestion_timestamp = db.Column(
                db.TIMESTAMP, default=datetime.now(timezone.utc))

        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "original_customer_id", "mobile_number", "pan_number",
            "aadhaar_number", "ucid_number", "loan_application_number",
            "reason_for_duplication", "master_customer_id",
            "ingestion_timestamp"
        ]
        writer.writerow(headers)

        duplicate_records = db.session.query(DuplicateCustomerRecord).all()

        for record in duplicate_records:
            row = [
                record.original_customer_id,
                record.mobile_number,
                record.pan_number,
                record.aadhaar_number,
                record.ucid_number,
                record.loan_application_number,
                record.reason_for_duplication,
                record.master_customer_id,
                record.ingestion_timestamp.isoformat()
                if record.ingestion_timestamp else ''
            ]
            writer.writerow(row)

        output.seek(0)
        return output

    def export_unique_data_csv(self) -> io.StringIO:
        """
        Generates a CSV file containing unique customer records after
        deduplication. (FR33)

        ASSUMPTION: The `customers` table itself contains the unique,
        deduplicated customer profiles.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "customer_id", "mobile_number", "pan_number", "aadhaar_number",
            "ucid_number", "loan_application_number", "dnd_flag", "segment",
            "created_at", "updated_at"
        ]
        writer.writerow(headers)

        unique_customers = db.session.query(Customer).all()

        for customer in unique_customers:
            row = [
                customer.customer_id,
                customer.mobile_number,
                customer.pan_number,
                customer.aadhaar_number,
                customer.ucid_number,
                customer.loan_application_number,
                "Yes" if customer.dnd_flag else "No",
                customer.segment,
                customer.created_at.isoformat()
                if customer.created_at else '',
                customer.updated_at.isoformat()
                if customer.updated_at else ''
            ]
            writer.writerow(row)

        output.seek(0)
        return output

    def export_error_excel_file(self) -> io.BytesIO:
        """
        Generates an Excel file detailing errors from data ingestion processes.
        (FR34)
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ingestion Errors"

        headers = [
            "Log ID", "File Name", "Upload Timestamp", "Status",
            "Error Description"
        ]
        sheet.append(headers)

        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD",
                                  end_color="4F81BD",
                                  fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        error_logs = db.session.query(IngestionLog).filter(
            IngestionLog.status == 'FAILED'
        ).order_by(IngestionLog.upload_timestamp.desc()).all()

        for log in error_logs:
            row = [
                log.log_id,
                log.file_name,
                log.upload_timestamp.isoformat()
                if log.upload_timestamp else '',
                log.status,
                log.error_description
            ]
            sheet.append(row)

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            # Get the column name (e.g., 'A', 'B')
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_name].width = adjusted_width

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output